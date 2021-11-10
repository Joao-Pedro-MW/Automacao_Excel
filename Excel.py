from SMTP import enviar_email
from tkinter import messagebox
from openpyxl import *

#abre o arquivo
nomes,emails = list(),list()
messagebox.showwarning("AVISO!", "PROCESSANDO DADOS, AGUARDE...")
planilha = load_workbook("Data.xlsx")
planilha = planilha["Nomes"]
#este loop acessa os valores na primeira coluna
for row in planilha.iter_rows(min_row=2, min_col=1, max_col=1,values_only=True):
    for cell in row: 
        nomes.append(cell)
#este loop acessa os valores na segunda coluna
for col in planilha.iter_cols(min_row=2, min_col=2, max_col=2,values_only=True):
    for cell in col:
        emails.append(cell)
#Gmail 
email_remetente = input("Informe o seu e-mail: ")
senha_remetente = input("Informe a sua senha de login: ")
assunto = str(input("Qual o assunto da mensagem a ser passada à todos os e-mails? "))
for n in range(len(nomes)):
    nome_do_destinatario = nomes[n]
    email_usuario = emails[n]
    print("Email sendo enviado para {user} em nome de {name}".format(user = email_usuario, name = nome_do_destinatario))
    enviar_email(email_remetente,email_usuario,senha_remetente,assunto,nome_do_destinatario)
